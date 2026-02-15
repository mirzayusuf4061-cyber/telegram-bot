import telebot
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os, math

# =========================
# Telegram Token (Render env variable orqali)
TOKEN = os.getenv("8173974361:AAER7WriV47EKKl5JjKaHRBflLdTR18Cv20")
bot = telebot.TeleBot(TOKEN)

# =========================
# Admin Telegram ID
ADMIN_ID = 123456789  # O'zingizning telegram ID bilan almashtiring

# =========================
# Javob kaliti
answer_keys = {
    "exam1": {
        "letters": "ABCDABCDABCDABCDABCDABCDABCDABCDABC",
        "written": ["12","5","ABC","100","7"]
    }
}

# =========================
# Rasch uchun natijalar
all_results = {}

# =========================
# Excel fayl nomi
FILE_NAME = "results.xlsx"

# =========================
# Excelga saqlash funksiyasi
def save_to_excel(user, score):
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["Ism", "User ID", "Ball", "Sana"])
        wb.save(FILE_NAME)

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    ws.append([
        user.full_name,
        user.id,
        score,
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ])
    wb.save(FILE_NAME)

# =========================
# Rasch uchun raw natijalar saqlash
def save_raw_result(user_id, letters, written):
    correct_letters = answer_keys["exam1"]["letters"]
    correct_written = answer_keys["exam1"]["written"]
    user_vector = []

    for i in range(35):
        user_vector.append(1 if letters[i].upper() == correct_letters[i].upper() else 0)
    for i in range(5):
        user_vector.append(1 if written[i].upper() == correct_written[i].upper() else 0)

    all_results[user_id] = user_vector

# =========================
# /check buyrug'i - faqat to'g'ri javoblar
@bot.message_handler(commands=['check'])
def check_exam(message):
    parts = message.text.split(maxsplit=2)
    if len(parts) < 3:
        bot.reply_to(message, "Format noto‘g‘ri. /check exam_id javoblar")
        return

    exam_id = parts[1]
    full_text = parts[2]

    letters_part = full_text[:35]
    written_part = full_text[35:].strip()
    written_answers = [x.strip().upper() for x in written_part.split(",")]

    correct_letters = answer_keys[exam_id]["letters"]
    correct_written = answer_keys[exam_id]["written"]

    # Faqat to'g'ri javoblar
    correct_list = []
    for i in range(35):
        if letters_part[i].upper() == correct_letters[i].upper():
            correct_list.append(str(i+1))
    for i in range(5):
        if written_answers[i].upper() == correct_written[i].upper():
            correct_list.append(str(35+i+1))

    result_text = "To‘g‘ri javoblar: " + ", ".join(correct_list)
    bot.reply_to(message, result_text)

    # Rasch uchun saqlash
    save_raw_result(message.from_user.id, letters_part, written_answers)

# =========================
# /rasch buyrug'i (Admin)
@bot.message_handler(commands=['rasch'])
def calculate_rasch(message):
    if message.from_user.id != ADMIN_ID:
        return

    for user_id, responses in all_results.items():
        correct = sum(responses)
        wrong = len(responses) - correct

        if wrong == 0: wrong = 0.5
        if correct == 0: correct = 0.5

        ability = round(math.log(correct / wrong), 3)
        bot.send_message(user_id,
            f"Rasch bo‘yicha umumiy natijangiz:\nAbility = {ability}"
        )

# =========================
# Botni ishga tushirish
bot.polling()
